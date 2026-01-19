import openpyxl
import csv
import os
import random
import string

def ConvertExcelToCSV(excelPath, outputCSVPath):
    try:
        wb = openpyxl.load_workbook(excelPath, data_only=True)
        sheet = wb.active
        
        hFirst = str(sheet.cell(row=2, column=2).value or "").strip()
        hLast = str(sheet.cell(row=2, column=3).value or "").strip()
        hEmail = str(sheet.cell(row=2, column=4).value or "").strip()

        if hFirst.lower() != "first" or hLast.lower() != "last" or hEmail.lower() != "email":
            return False, f"Invalid Template. Expected headers 'First', 'Last', 'Email' on Row 2 (Cols B,C,D). Found: '{hFirst}', '{hLast}', '{hEmail}'."

        rowsProcessed = 0
        with open(outputCSVPath, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["FirstName", "LastName", "Email", "Password"]) 

            for row_idx in range(3, sheet.max_row + 1):
                first = str(sheet.cell(row=row_idx, column=2).value or "").strip()
                last = str(sheet.cell(row=row_idx, column=3).value or "").strip()
                email = str(sheet.cell(row=row_idx, column=4).value or "").strip()

                if not first or not last or not email:
                    continue
                
                chars = string.ascii_letters + string.digits + "!@#$%"
                pwd = "".join(random.choices(chars, k=12))

                writer.writerow([first, last, email, pwd])
                rowsProcessed += 1

        return True, f"Successfully converted {rowsProcessed} users."

    except Exception as e:
        return False, f"Error: {str(e)}"